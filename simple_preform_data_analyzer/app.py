import os
import re
import threading
from collections import defaultdict
from datetime import datetime
from tkinter import Tk, Button, Label, Text, filedialog, StringVar, END, messagebox
from tkinter.ttk import Combobox

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill

import matplotlib.pyplot as plt


mapping = [
    ("spoolno2", 1, None, None),
    ("OTDR length", 9, None, None),
    ("Attenuation 1310 I/E", 5, None, None),
    ("Attenuation 1310 O/E", 6, None, None),
    ("Attenuation 1383 I/E", 73, None, None),
    ("Attenuation 1383 O/E", 74, None, None),
    ("Attenuation 1550 I/E", 7, None, None),
    ("Attenuation 1550 O/E", 8, None, None),
    ("Attenuation 1625 I/E", 75, None, None),
    ("Attenuation 1625 O/E", 76, None, None),
    ("MFD 1310nm I/E", 12, None, None),
    ("MFD 1310nm O/E", 13, None, None),
    ("", None, None, None),
    ("", None, None, None),
    ("", None, None, None),
    ("", None, None, None),
    ("", None, None, None),
    ("", None, None, None),
    ("Cutoff 2m I/E", 14, None, None),
    ("Cutoff 2m O/E", 15, None, None),
    ("Cutoff 22m", 24, None, None),
    ("delta 2m-22m", None, "delta", None),
    ("Mac value", None, "mac", None),
    ("Clad Dia. I/E", 16, None, None),
    ("Clad Dia. O/E", 17, None, None),
    ("Clad Ovality I/E", 18, None, None),
    ("Clad Ovality O/E", 19, None, None),
    ("Core Ovality I/E", 20, None, None),
    ("Core Ovality O/E", 21, None, None),
    ("ECC I/E", 22, None, None),
    ("ECC O/E", 23, None, None),
    ("Zero Dispersion Wave.", 30, None, None),
    ("dispslope at ZDW", 31, None, None),
    ("Dispersion 1285", 32, None, None),
    ("Dispersion 1290", 33, None, None),
    ("Dispersion 1330", 34, None, None),
    ("Dispersion 1550", 35, None, None),
    ("", None, None, None),
    ("PMD", 37, None, None),
    ("R7.5mm 1t 1550", 26, "scale", 0.1),
    ("R7.5mm 1t 1625", 69, "scale", 0.1),
    ("R10mm 1t 1550", 70, "scale", 0.1),
    ("R10mm 1t 1625", 71, "scale", 0.1),
    ("R15mm 10t 1550", 81, "scale", 0.5),
    ("R15mm 10t 1625", 82, "scale", 0.5),
]

red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")


def safe_sheet_name(name):
    name = str(name).strip()
    name = re.sub(r'[\[\]\:\*\?\/\\]', "_", name)

    if name == "" or name.lower() == "nan" or name == "None":
        name = "Blank"

    return name[:31]


def safe_file_name(name):
    name = str(name).strip()
    name = re.sub(r'[\\/:*?"<>|]', "_", name)

    if name == "" or name.lower() == "nan" or name == "None":
        name = "Blank"

    return name


def is_number(value):
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def replace_zero_with_blank(ws):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == 0 or cell.value == "0":
                cell.value = None


def replace_zero_in_workbook(wb):
    for ws in wb.worksheets:
        replace_zero_with_blank(ws)


def delete_rows_if_E_last_digit_not_zero(ws):
    rows_to_delete = []

    for row_idx in range(2, ws.max_row + 1):
        value = ws.cell(row=row_idx, column=5).value

        if value is None:
            continue

        s = str(value).strip()

        if s == "":
            continue

        last_char = s[-1]

        if last_char.isdigit() and last_char != "0":
            rows_to_delete.append(row_idx)

    for row_idx in reversed(rows_to_delete):
        ws.delete_rows(row_idx)


def get_break_count(spoolno):
    """
    A=0, B=1, C=2, D=3 ...
    spoolno 끝에서 5, 6, 7번째 문자 중 가장 큰 값 사용
    """
    if spoolno is None:
        return None

    s = str(spoolno).upper()
    break_values = []

    for pos in [5, 6, 7]:
        if len(s) >= pos:
            ch = s[-pos]

            if "A" <= ch <= "Z":
                break_values.append(ord(ch) - ord("A"))

    return max(break_values) if break_values else None


def check_range(value, low, high):
    if value is None:
        return False

    if not is_number(value):
        return False

    return value < low or value > high


class ExcelApp:
    def __init__(self, root):
        self.root = root
        self.root.title("LS Excel Preform")
        self.root.geometry("850x680")

        self.input_file = None
        self.output_dir = None
        self.created_files = []
        self.log_file = None

        self.file_label_var = StringVar(value="선택된 파일 없음")

        Label(root, text="LS Excel Preform", font=("Arial", 20, "bold")).pack(pady=10)

        Button(root, text="alls.xlsx 파일 선택", width=35, command=self.select_file).pack(pady=5)

        Label(root, textvariable=self.file_label_var, fg="blue").pack(pady=5)

        Button(
            root,
            text="1단계 실행: 그룹별 xlsx 생성",
            width=45,
            command=self.run_step1_thread
        ).pack(pady=8)

        Label(root, text="어떤 값에 관심 있으세요?").pack(pady=5)

        self.combo = Combobox(root, state="readonly", width=40)
        self.combo.pack(pady=5)

        Button(
            root,
            text="2단계 실행: 선택 파일 Report 생성",
            width=45,
            command=self.run_step2_thread
        ).pack(pady=8)

        Button(
            root,
            text="3단계 실행: 월별 결과값 보기 및 그래프 저장",
            width=45,
            command=self.run_step3_thread
        ).pack(pady=8)

        Label(root, text="작업 진행 상태").pack(pady=5)

        self.log_box = Text(root, height=23, width=105)
        self.log_box.pack(padx=10, pady=10)

    def log(self, msg):
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_text = f"[{now}] {msg}"

        self.log_box.insert(END, log_text + "\n")
        self.log_box.see(END)
        self.root.update_idletasks()

        if self.log_file:
            with open(self.log_file, "a", encoding="utf-8") as f:
                f.write(log_text + "\n")

    def select_file(self):
        file_path = filedialog.askopenfilename(
            title="alls.xlsx 파일 선택",
            filetypes=[("Excel files", "*.xlsx")]
        )

        if file_path:
            self.input_file = file_path
            self.output_dir = os.path.dirname(file_path)
            self.file_label_var.set(file_path)

            self.log_file = os.path.join(self.output_dir, "run_log.txt")

            with open(self.log_file, "w", encoding="utf-8") as f:
                f.write("LS Excel Preform 실행 로그\n")
                f.write(f"로그 시작 시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")

            self.log(f"파일 선택 완료: {file_path}")

    def run_step1_thread(self):
        threading.Thread(target=self.run_step1, daemon=True).start()

    def run_step2_thread(self):
        threading.Thread(target=self.run_step2, daemon=True).start()

    def run_step3_thread(self):
        threading.Thread(target=self.run_step3, daemon=True).start()

    def calc_column_avg(self, ws, col_idx):
        nums = []

        for row_idx in range(2, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value

            if is_number(value):
                nums.append(value)

        return sum(nums) / len(nums) if nums else None

    def run_step1(self):
        if not self.input_file:
            self.log("먼저 alls.xlsx 파일을 선택해주세요.")
            return

        try:
            self.log("alls.xlsx 파일 로드 실행중입니다.")
            wb = load_workbook(self.input_file)
            source_ws = wb.active

            self.log("alls.xlsx 내 0 값을 공백으로 변환 실행중입니다.")
            replace_zero_in_workbook(wb)

            self.log("원본 데이터 읽기 실행중입니다.")
            rows = list(source_ws.iter_rows(values_only=True))

            if len(rows) < 2:
                self.log("데이터가 부족합니다.")
                return

            header = rows[0]
            data_rows = rows[1:]

            self.log("E열 기준 앞 2글자 그룹화 실행중입니다.")
            group_by_2char = defaultdict(list)

            for row in data_rows:
                value = row[4]
                group_name = "Blank" if value is None else str(value)[:2]
                group_by_2char[group_name].append(row)

            self.log("alls.xlsx 내부 그룹 시트 생성 실행중입니다.")

            for group_name, group_rows in group_by_2char.items():
                sheet_name = safe_sheet_name(group_name)

                if sheet_name in wb.sheetnames:
                    del wb[sheet_name]

                ws = wb.create_sheet(sheet_name)
                ws.append(header)

                for row in group_rows:
                    ws.append(row)

            wb.save(self.input_file)

            self.log("그룹별 xlsx 파일 생성 실행중입니다.")
            self.created_files = []

            for group_name in group_by_2char.keys():
                group_sheet_name = safe_sheet_name(group_name)
                group_ws = wb[group_sheet_name]

                group_rows_all = list(group_ws.iter_rows(values_only=True))
                group_header = group_rows_all[0]
                group_data_rows = group_rows_all[1:]

                sub_groups = defaultdict(list)

                for row in group_data_rows:
                    value = row[3]
                    sub_sheet_name = "Blank" if value is None else str(value)
                    sub_groups[sub_sheet_name].append(row)

                new_wb = Workbook()
                new_wb.remove(new_wb.active)

                for sub_name, sub_rows in sub_groups.items():
                    ws_name = safe_sheet_name(sub_name)
                    new_ws = new_wb.create_sheet(ws_name)

                    new_ws.append(group_header)

                    for row in sub_rows:
                        new_ws.append(row)

                output_file = os.path.join(
                    self.output_dir,
                    f"{safe_file_name(group_name)}.xlsx"
                )

                new_wb.save(output_file)
                self.created_files.append(output_file)

            options = [
                os.path.splitext(os.path.basename(f))[0]
                for f in self.created_files
            ]

            self.combo["values"] = options

            if options:
                self.combo.current(0)

            self.log(f"1단계 완료. 생성된 파일: {options}")
            self.log(f"{options} 중 고르세요.")

        except Exception as e:
            self.log(f"오류 발생: {e}")

    def run_step2(self):
        selected = self.combo.get()

        if not selected:
            self.log("먼저 관심 값을 선택해주세요.")
            return

        selected_file = os.path.join(self.output_dir, f"{selected}.xlsx")

        if not os.path.exists(selected_file):
            self.log(f"{selected_file} 파일이 없습니다.")
            return

        start_time = datetime.now()
        error_logs = []

        try:
            self.log(f"{selected_file} 파일 로드 실행중입니다.")
            target_wb = load_workbook(selected_file)

            self.log(f"{selected_file} 내 E열 조건 필터링 실행중입니다.")
            for ws in target_wb.worksheets:
                delete_rows_if_E_last_digit_not_zero(ws)

            self.log(f"{selected_file} 내 0 값을 공백으로 변환 실행중입니다.")
            replace_zero_in_workbook(target_wb)

            self.log("시트별 평균값 및 단선 횟수 계산 실행중입니다.")
            sheet_result_info = []

            for ws in target_wb.worksheets:
                max_row = ws.max_row

                if max_row < 2:
                    continue

                break_counts = []

                for row_idx in range(2, max_row + 1):
                    spoolno = ws.cell(row=row_idx, column=5).value
                    break_count = get_break_count(spoolno)

                    if break_count is not None:
                        break_counts.append(break_count)

                final_break_count = max(break_counts) if break_counts else None

                last_values_A_to_E = [
                    ws.cell(row=max_row, column=col).value
                    for col in range(1, 6)
                ]

                avg_values = []

                for col in range(6, 86):
                    nums = []

                    for row_idx in range(2, max_row + 1):
                        value = ws.cell(row=row_idx, column=col).value

                        if is_number(value):
                            nums.append(value)

                    avg_values.append(sum(nums) / len(nums) if nums else None)

                avg_row_idx = max_row + 2

                for col_idx, value in enumerate(last_values_A_to_E, start=1):
                    ws.cell(row=avg_row_idx, column=col_idx).value = value

                for col_idx, avg_value in enumerate(avg_values, start=6):
                    ws.cell(row=avg_row_idx, column=col_idx).value = avg_value

                ws.cell(row=avg_row_idx - 1, column=86).value = "단선 횟수"
                ws.cell(row=avg_row_idx, column=86).value = final_break_count

                sheet_result_info.append({
                    "sheet_name": ws.title,
                    "avg_row_idx": avg_row_idx,
                    "break_count": final_break_count
                })

            self.log(f"{selected}.xlsx 저장 실행중입니다.")
            target_wb.save(selected_file)

            self.log("Report 파일 작성 실행중입니다.")
            report_wb = Workbook()
            report_ws = report_wb.active
            report_ws.title = "Report"

            report_header = ["Sheet Name"] + [item[0] for item in mapping] + ["단선 횟수", "hdate"]
            report_ws.append(report_header)

            for info in sheet_result_info:
                ws = target_wb[info["sheet_name"]]
                avg_row_idx = info["avg_row_idx"]

                report_row = [ws.title]

                for name, col_idx, op, param in mapping:

                    if col_idx is None and op is None:
                        report_row.append(None)

                    elif col_idx is not None:
                        value = ws.cell(row=avg_row_idx, column=col_idx + 1).value

                        if op == "scale" and value is not None and is_number(value):
                            value = value * param

                        report_row.append(value)

                    elif op == "delta":
                        cutoff_2m = ws.cell(row=avg_row_idx, column=15).value
                        cutoff_22m = ws.cell(row=avg_row_idx, column=25).value

                        if is_number(cutoff_2m) and is_number(cutoff_22m):
                            report_row.append(cutoff_2m - cutoff_22m)
                        else:
                            report_row.append(None)

                    elif op == "mac":
                        report_row.append(None)

                report_row.append(info["break_count"])

                hdate_value = ws.cell(row=avg_row_idx, column=1).value
                report_row.append(hdate_value)

                report_ws.append(report_row)

                current_row = report_ws.max_row
                a_value = report_ws.cell(row=current_row, column=1).value
                row_error_messages = []

                val_L = report_ws.cell(row=current_row, column=12).value
                val_M = report_ws.cell(row=current_row, column=13).value

                if check_range(val_L, 8.8, 9.2) or check_range(val_M, 8.8, 9.2):
                    row_error_messages.append(f"{a_value}->MFD 오류 발생")

                val_W = report_ws.cell(row=current_row, column=23).value

                if is_number(val_W) and val_W < 0:
                    row_error_messages.append(f"{a_value}->Cutoff delta2m-22m 오류 발생")

                val_Y = report_ws.cell(row=current_row, column=25).value
                val_Z = report_ws.cell(row=current_row, column=26).value

                if check_range(val_Y, 124.3, 125.7) or check_range(val_Z, 124.3, 125.7):
                    row_error_messages.append(f"{a_value}->Clad Dia. 오류 발생")

                val_AH = report_ws.cell(row=current_row, column=34).value

                if check_range(val_AH, 0.073, 0.09):
                    row_error_messages.append(f"{a_value}->disp slope at ZDW 오류 발생")

                if row_error_messages:
                    for col in range(1, report_ws.max_column + 1):
                        report_ws.cell(row=current_row, column=col).fill = red_fill

                    for msg in row_error_messages:
                        now = datetime.now()
                        self.log(msg)
                        error_logs.append({
                            "time": now,
                            "msg": msg
                        })

            report_file = os.path.join(self.output_dir, f"{selected}_report.xlsx")
            report_wb.save(report_file)

            if error_logs:
                error_file = os.path.join(self.output_dir, "error.txt")

                with open(error_file, "w", encoding="utf-8") as f:
                    f.write(f"[실행 시작 시간] {start_time.strftime('%Y-%m-%d %H:%M:%S')}\n\n")

                    for err in error_logs:
                        f.write(f"[오류 발생 시간] {err['time'].strftime('%Y-%m-%d %H:%M:%S')}\n")
                        f.write(f"{err['msg']}\n\n")

                self.log(f"error.txt 파일 생성 완료: {error_file}")
            else:
                self.log("품질 오류 없음.")

            self.log("전체 작업 완료.")
            self.log(f"Report 생성 완료: {report_file}")

        except Exception as e:
            self.log(f"오류 발생: {e}")

    def run_step3(self):
        selected = self.combo.get()

        if not selected:
            self.log("먼저 관심 값을 선택해주세요.")
            return

        report_file = os.path.join(self.output_dir, f"{selected}_report.xlsx")

        if not os.path.exists(report_file):
            self.log(f"{report_file} 파일이 없습니다. 먼저 2단계를 실행해주세요.")
            return

        try:
            self.log(f"{selected}_report.xlsx 월별 결과값 분석 실행중입니다.")

            wb = load_workbook(report_file)
            report_ws = wb["Report"]

            hdate_col = report_ws.max_column
            monthly_rows = {}

            for row_idx in range(2, report_ws.max_row + 1):
                hdate = report_ws.cell(row=row_idx, column=hdate_col).value

                if hdate is None:
                    continue

                hdate_str = str(hdate).strip()
                parts = hdate_str.split(".")

                if len(parts) >= 2:
                    month_key = f"{parts[0]}.{parts[1]}"
                else:
                    continue

                if month_key not in monthly_rows:
                    monthly_rows[month_key] = []

                monthly_rows[month_key].append(row_idx)

            if not monthly_rows:
                self.log("월별로 분류할 hdate 값이 없습니다.")
                return

            monthly_summary = []

            for month_key in sorted(monthly_rows.keys()):
                row_indices = monthly_rows[month_key]
                sheet_name = safe_sheet_name(month_key)

                if sheet_name in wb.sheetnames:
                    del wb[sheet_name]

                month_ws = wb.create_sheet(sheet_name)

                for col in range(1, report_ws.max_column + 1):
                    month_ws.cell(row=1, column=col).value = report_ws.cell(row=1, column=col).value

                new_row = 2

                for row_idx in row_indices:
                    for col in range(1, report_ws.max_column + 1):
                        month_ws.cell(row=new_row, column=col).value = report_ws.cell(row=row_idx, column=col).value
                    new_row += 1

                avg_H = self.calc_column_avg(month_ws, 8)
                avg_I = self.calc_column_avg(month_ws, 9)
                avg_L = self.calc_column_avg(month_ws, 12)
                avg_M = self.calc_column_avg(month_ws, 13)

                monthly_summary.append({
                    "month": month_key,
                    "att_ie": avg_H,
                    "att_oe": avg_I,
                    "mfd_ie": avg_L,
                    "mfd_oe": avg_M,
                })

                self.log(f"1. {month_key}의 Att.1550nm I/E - H열 평균값: {avg_H}")
                self.log(f"2. {month_key}의 Att.1550nm O/E - I열 평균값: {avg_I}")
                self.log(f"3. {month_key}의 MFD 1310nm I/E - L열 평균값: {avg_L}")
                self.log(f"4. {month_key}의 MFD 1310nm O/E - M열 평균값: {avg_M}")
                self.log("")

            wb.save(report_file)

            months = [item["month"] for item in monthly_summary]

            graph_items = [
                ("att1550nm_IE", "Att.1550nm I/E", "att_ie"),
                ("att1550nm_OE", "Att.1550nm O/E", "att_oe"),
                ("mfd1310nm_IE", "MFD 1310nm I/E", "mfd_ie"),
                ("mfd1310nm_OE", "MFD 1310nm O/E", "mfd_oe"),
            ]

            for file_name, title, key in graph_items:
                y_values = [item[key] for item in monthly_summary]

                plt.figure(figsize=(10, 5))
                plt.plot(months, y_values, marker="o")
                plt.title(title)
                plt.xlabel("Month")
                plt.ylabel(title)
                plt.xticks(rotation=45)
                plt.grid(True)
                plt.tight_layout()

                png_file = os.path.join(self.output_dir, f"{selected}_{file_name}.png")
                plt.savefig(png_file, dpi=300)
                plt.close()

                self.log(f"그래프 저장 완료: {png_file}")

            self.log(f"{selected}_report.xlsx 월별 시트 생성 및 평균값 출력 완료.")

        except Exception as e:
            self.log(f"오류 발생: {e}")


if __name__ == "__main__":
    root = Tk()

    messagebox.showinfo(
        "LS Excel Preform 안내",
        "안녕하세요?\n\n"
        "저는 LS전선 통신시스템연구그룹 김희서 연구원입니다.\n\n"
        "파이썬을 기반으로 광섬유 품질 이력 값 정리를 위한 분석 툴입니다.\n\n"
        "관련 문의 사항은\n"
        "hkim17@lscns.com\n"
        "으로 문의주시길 바랍니다."
    )

    app = ExcelApp(root)
    root.mainloop()